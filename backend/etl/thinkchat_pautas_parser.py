"""
thinkchat_pautas_parser.py — Parsea el Excel de pautas ThinkChat.
Mapea Linea (ODO/MPP/Estetica) -> enterprise_id (1/2/5).
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl

logger = logging.getLogger("crm.etl.thinkchat_pautas_parser")

# Mapeo de valores de Línea en el Excel a enterprise_id del DW
# Cobertura flexible para tolerar variaciones: mayus, minus, espacios
LINEA_TO_ENTERPRISE = {
    "odontologia": 1,
    "odo": 1,
    "medicina prepaga": 2,
    "mpp": 2,
    "med. prepaga": 2,
    "medicina estetica": 5,
    "medicina est\u00e9tica": 5,
    "estetica": 5,
    "est\u00e9tica": 5,
    "est": 5,
    "epem": 5,
    "grupo epem": 5,
    # Emergencias: source 6 en ThinkChat pero mapeado a enterprise_id 4 en EPEM
    "emergencias": 4,
    "eme": 4,
    "med. emergencias": 4,
    "medicina emergencias": 4,
}


def _norm(s: str) -> str:
    return (s or "").strip().lower()


def _to_str(val) -> str:
    """Convierte cualquier valor a string vacio si es None o no es texto."""
    if val is None:
        return ""
    if isinstance(val, str):
        return val.strip()
    return str(val).strip()


def parse_pautas_xlsx(path: Path) -> list[dict]:
    """
    Lee el Excel y devuelve una lista de dicts normalizados:
    [{
      'phone': '595991816836',
      'fullname': 'Cristian V',
      'enterprise_id': 1,
      'ad_id': '120210976867940147',
      'campaign_name': 'Agenda una cita',
      'canal': 'WhatsApp',
      'ultimo_estado': '...',
      'agente': '...',
      'sucursal': '...',
      'fecha_ingreso': datetime(2025,1,1,0,22),
      'linea_raw': 'Odontologia',
    }, ...]
    Devuelve [] si el archivo no existe o no es xlsx valido.
    """
    if not path.exists():
        logger.error(f"Excel no existe: {path}")
        return []

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"No se pudo abrir {path}: {e}")
        return []

    if "ThinkChat" not in wb.sheetnames:
        logger.error(f"Sheet 'ThinkChat' no esta en el Excel. Sheets: {wb.sheetnames}")
        return []

    ws = wb["ThinkChat"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = list(rows[0])
    col = {name: i for i, name in enumerate(header)}

    # Validar columnas requeridas
    required = ["L\u00ednea", "Linea", "Fecha Ingreso", "Contacto", "Meta AD ID", "Canal", "\u00daltimo Estado"]
    missing = [c for c in required if c not in col]
    if missing:
        logger.error(f"Columnas faltantes en Excel: {missing}")
        return []

    leads = []
    skipped = {"no_phone": 0, "no_linea": 0, "bad_date": 0, "unknown_linea": 0}

    for r in rows[1:]:
        if not r or all(v is None for v in r):
            continue

        linea_raw = r[col["L\u00ednea"]]
        linea_key = _norm(str(linea_raw)) if linea_raw else ""
        enterprise_id = LINEA_TO_ENTERPRISE.get(linea_key)

        if not linea_raw:
            skipped["no_linea"] += 1
            continue
        if enterprise_id is None:
            skipped["unknown_linea"] += 1
            continue

        # Telefono: columna 'Linea' trae el numero (formato 595...)
        phone_raw = r[col["Linea"]]
        if not phone_raw:
            skipped["no_phone"] += 1
            continue
        phone = str(phone_raw).strip()
        if len(phone) < 9:
            skipped["no_phone"] += 1
            continue

        # Fecha
        fecha_raw = r[col["Fecha Ingreso"]]
        fecha_ingreso = None
        if fecha_raw:
            if isinstance(fecha_raw, datetime):
                fecha_ingreso = fecha_raw
            else:
                # Formato tipico: "01/01/2025 00:22"
                for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
                    try:
                        fecha_ingreso = datetime.strptime(str(fecha_raw), fmt)
                        break
                    except (ValueError, TypeError):
                        continue
        if fecha_ingreso is None:
            skipped["bad_date"] += 1
            continue

        leads.append({
            "phone": phone,
            "fullname": _to_str(r[col["Contacto"]]) or "Sin nombre",
            "enterprise_id": enterprise_id,
            "ad_id": _to_str(r[col["Meta AD ID"]]),
            "campaign_name": _to_str(r[col["Pauta Titulo"]]) if "Pauta Titulo" in col else "",
            "canal": _to_str(r[col["Canal"]]),
            "ultimo_estado": _to_str(r[col["\u00daltimo Estado"]]),
            "agente": _to_str(r[col["Agente"]]) if "Agente" in col else "",
            "sucursal": _to_str(r[col["Sucursal"]]) if "Sucursal" in col else "",
            "fecha_ingreso": fecha_ingreso,
            "linea_raw": str(linea_raw),
        })

    logger.info(
        f"Parsed {len(leads)} leads from {path.name}. "
        f"Skipped: {skipped}"
    )
    return leads
